import os
import json
import datetime
import logging
import requests
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional
import zipfile
from openpyxl import load_workbook
import xlrd
import threading
import time
import sys

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class SimpleOptionsProcessor:
    """Simple and reliable options data processor"""

    def __init__(self):
        self.base_path = Path.cwd()
        self.old_version_path = self.base_path / "old version"
        self.new_version_path = self.base_path / "new version"
        self.output_path = self._get_output_path()

        # Setup directories
        self.old_version_path.mkdir(exist_ok=True)
        self.new_version_path.mkdir(exist_ok=True)
        self.output_path.mkdir(parents=True, exist_ok=True)

        # Configuration
        self.currencies = {
            'EUR': {'id': '58', 'coefficient': 10000, 'option_type': 'OPTION TYPE: Monthly Options'},
            'GBP': {'id': '42', 'coefficient': 1000, 'option_type': 'OPTION TYPE: Monthly Options'},
            'AUD': {'id': '37', 'coefficient': 10000, 'option_type': 'OPTION TYPE: Monthly Options'},
            'CAD': {'id': '48', 'coefficient': 10000000, 'option_type': 'OPTION TYPE: Monthly Options'},
            'JPY': {'id': '69', 'coefficient': 1000000, 'option_type': 'OPTION TYPE: Monthly Options'},
            'XAU': {'id': '437', 'coefficient': 1, 'option_type': 'OPTION TYPE: American Options'},
            'XAG': {'id': '458', 'coefficient': 100, 'option_type': 'OPTION TYPE: American Options'}
        }

        # Create session for downloads
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

    def _get_output_path(self) -> Path:
        """Get output path for JSON files"""
        try:
            appdata = os.environ.get('APPDATA')
            if appdata:
                mt5_path = Path(appdata) / 'MetaQuotes' / 'Terminal' / 'Common' / 'Files'
                if mt5_path.exists():
                    return mt5_path
        except:
            pass
        return self.base_path / 'output'

    def show_info(self):
        """Show directory information"""
        print("\n" + "=" * 60)
        print("📁 РАБОЧИЕ ПАПКИ")
        print("=" * 60)
        print(f"📂 Базовая папка: {self.base_path}")
        print(f"📂 Временные XLS: {self.old_version_path}")
        print(f"📂 XLSX файлы: {self.new_version_path}")
        print(f"📂 JSON результаты: {self.output_path}")
        print("=" * 60)

    def get_trading_date(self, days_back: int = 2) -> str:
        """Get trading date (skip weekends)"""
        date = datetime.datetime.now()
        while days_back > 0:
            date -= datetime.timedelta(days=1)
            if date.weekday() < 5:  # Skip weekends
                days_back -= 1
        return date.strftime("%Y%m%d")

    def download_cme_files(self) -> bool:
        """Download CME options data files"""
        try:
            # Check if files already exist
            existing_files = list(self.old_version_path.glob('*.xls'))
            if existing_files:
                logger.info("📁 XLS файлы уже существуют")
                return True

            trade_date = self.get_trading_date()
            url = 'https://www.cmegroup.com/CmeWS/exp/voiProductDetailsViewExport.ctl'

            logger.info(f"📥 Загрузка данных CME за {trade_date}")

            for currency, info in self.currencies.items():
                try:
                    logger.info(f"⬇️ Загрузка {currency}...")

                    params = {
                        'media': 'xls',
                        'tradeDate': trade_date,
                        'reportType': 'P',
                        'productId': info['id']
                    }

                    response = self.session.get(url, params=params, timeout=30)
                    response.raise_for_status()

                    file_path = self.old_version_path / f'{currency}.xls'
                    with open(file_path, 'wb') as f:
                        f.write(response.content)

                    logger.info(f"✅ {currency} загружен")
                except Exception as e:
                    logger.error(f"❌ Ошибка загрузки {currency}: {e}")
                    return False

            logger.info("✅ Все файлы CME загружены")
            return True

        except Exception as e:
            logger.error(f"❌ Общая ошибка загрузки CME: {e}")
            return False

    def download_cftc_simple(self) -> bool:
        """Simple CFTC download"""
        try:
            year = datetime.datetime.now().year
            url = f"https://www.cftc.gov/files/dea/history/fut_fin_xls_{year}.zip"
            zip_path = f"fut_fin_{year}.zip"

            logger.info(f"📥 Загрузка CFTC данных за {year} год...")

            response = self.session.get(url, timeout=300, stream=True)
            response.raise_for_status()

            total_size = int(response.headers.get('content-length', 0))
            logger.info(f"📦 Размер файла: {total_size / (1024 * 1024):.1f} MB")

            downloaded = 0
            with open(zip_path, "wb") as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)

                        if total_size > 0:
                            progress = (downloaded / total_size) * 100
                            if progress % 25 < 0.8:  # Show every 25%
                                logger.info(f"📊 {progress:.0f}%")

            # Extract
            if zipfile.is_zipfile(zip_path):
                with zipfile.ZipFile(zip_path, 'r') as z:
                    z.extractall()
                os.remove(zip_path)
                logger.info("✅ CFTC данные загружены")
                return True
            else:
                logger.error("❌ Поврежденный файл")
                return False

        except Exception as e:
            logger.error(f"❌ Ошибка загрузки CFTC: {e}")
            return False

    def convert_files(self) -> bool:
        """Convert XLS to XLSX"""
        try:
            xls_files = list(self.old_version_path.glob('*.xls'))
            if not xls_files:
                logger.warning("⚠️ Нет XLS файлов для конвертации")
                return False

            logger.info(f"🔄 Конвертация {len(xls_files)} файлов...")

            for file_path in xls_files:
                try:
                    logger.info(f"🔄 {file_path.name}")

                    # Read XLS
                    df = pd.read_excel(file_path, engine='xlrd')

                    # Save as XLSX
                    output_path = self.new_version_path / f"{file_path.stem}.xlsx"
                    df.to_excel(output_path, index=False, engine='openpyxl')

                    # Remove original
                    file_path.unlink()

                    logger.info(f"✅ {file_path.name} → {output_path.name}")

                except Exception as e:
                    logger.error(f"❌ Ошибка конвертации {file_path.name}: {e}")
                    continue

            logger.info("✅ Конвертация завершена")
            return True

        except Exception as e:
            logger.error(f"❌ Общая ошибка конвертации: {e}")
            return False

    def get_close_prices(self) -> Optional[List[float]]:
        """Get close prices from MT5 using your exact function"""
        try:
            import MetaTrader5 as mt5

            logger.info("🔌 Подключение к MT5...")

            # Your exact working code
            mt5.initialize()
            login = 500289067
            password = '0vC!VxNj'
            server = 'ForexClub-MT5 Demo Server'
            mt5.login(login, password, server)
            symbols = ["EURUSD", "GBPUSD", "AUDUSD", "USDCAD", "USDJPY", "XAUUSD", "XAGUSD"]

            # Get prices
            closed_price = [mt5.copy_rates_from_pos(i, mt5.TIMEFRAME_D1, 1, 1)[0][1] for i in symbols]
            mt5.shutdown()

            # Your exact transformations
            closed_price[0] = round(closed_price[0] * 10000) /10000
            closed_price[1] = round(closed_price[1] * 1000) /1000
            closed_price[2] = round(closed_price[2] * 10000) /10000
            closed_price[3] = round(closed_price[3] * 10000) / 10000
            closed_price[4] = round(closed_price[4] * 1000000) / 1000000
            closed_price[5] = round(closed_price[5])
            closed_price[6] = round(closed_price[6])

            logger.info("✅ Цены получены из MT5:")
            for symbol, price in zip(symbols, closed_price):
                logger.info(f"   {symbol}: {price}")

            return closed_price

        except ImportError:
            logger.error("❌ Модуль MetaTrader5 не установлен")
            return None
        except Exception as e:
            logger.error(f"❌ Ошибка получения цен: {e}")
            return None

    def get_manual_prices(self) -> Optional[List[float]]:
        """Get prices through manual input"""
        try:
            print("\n" + "=" * 50)
            print("✏️  РУЧНОЙ ВВОД ЦЕН")
            print("=" * 50)

            defaults = [
                ("EURUSD", "1.0850"),
                ("GBPUSD", "1.2650"),
                ("AUDUSD", "0.6750"),
                ("USDCAD", "1.3650"),
                ("USDJPY", "148.50"),
                ("XAUUSD", "2050.00"),
                ("XAGUSD", "24.50")
            ]

            raw_prices = []

            for symbol, default in defaults:
                while True:
                    try:
                        user_input = input(f"📈 {symbol} (по умолчанию {default}): ").strip()
                        price = float(user_input) if user_input else float(default)

                        if price <= 0:
                            print("   ❌ Цена должна быть больше нуля")
                            continue

                        raw_prices.append(price)
                        print(f"   ✅ {symbol}: {price}")
                        break

                    except ValueError:
                        print("   ❌ Введите корректное число")
                        continue

            # Transform prices
            transformed = [
                round(raw_prices[0] * 10000),  # EURUSD
                round(raw_prices[1] * 1000),  # GBPUSD
                round(raw_prices[2] * 10000),  # AUDUSD
                round(1 / raw_prices[3] * 10000),  # USDCAD
                round(1 / raw_prices[4] * 1000000),  # USDJPY
                round(raw_prices[5]),  # XAUUSD
                round(raw_prices[6] * 100)  # XAGUSD
            ]

            logger.info("✅ Цены введены вручную")
            return transformed

        except KeyboardInterrupt:
            logger.info("🛑 Ввод отменен")
            return None
        except Exception as e:
            logger.error(f"❌ Ошибка ввода: {e}")
            return None

    def process_currency(self, currency: str, close_price: float) -> bool:
        """Process options data for currency"""
        try:
            currency_info = self.currencies[currency]
            file_path = self.new_version_path / f"{currency}.xlsx"

            if not file_path.exists():
                logger.error(f"❌ Файл не найден: {file_path}")
                return False

            logger.info(f"📊 Обработка {currency}...")

            # Load workbook
            wb = load_workbook(str(file_path))
            ws = wb.active

            # Find data ranges
            option_type = currency_info['option_type']
            ranges = self._find_ranges(ws, option_type)

            if not ranges:
                logger.error(f"❌ Не найдены диапазоны для {currency}")
                return False

            # Extract data
            call_data = self._extract_data(ws, ranges['call'])
            put_data = self._extract_data(ws, ranges['put'])

            if not call_data['strike'] or not put_data['strike']:
                logger.error(f"❌ Нет данных для {currency}")
                return False

            # Calculate metrics
            coefficient = currency_info['coefficient']
            sip_metrics = self._calculate_sip(call_data, put_data, close_price, coefficient, currency)
            fob_metrics = self._calculate_fob(call_data, put_data, close_price)
            strike_data = self._format_strikes(call_data, put_data, coefficient, currency)

            # Create result
            result = {
                'strike': strike_data,
                'sip': sip_metrics,
                'fob': fob_metrics
            }

            # Save JSON
            self._save_json(result, currency)

            logger.info(f"✅ {currency} обработан")
            return True

        except Exception as e:
            logger.error(f"❌ Ошибка обработки {currency}: {e}")
            return False

    def _find_ranges(self, ws, option_type: str) -> Optional[Dict]:
        """Find data ranges in worksheet"""
        try:
            ranges = {'call': {}, 'put': {}}
            call_found = False

            for row in ws.iter_rows():
                for cell in row:
                    if not cell.value:
                        continue

                    cell_value = str(cell.value).strip()

                    if cell_value == option_type and not call_found:
                        ranges['call']['start'] = cell.row + 3
                        call_found = True
                    elif call_found and cell_value == 'TOTALS' and 'stop' not in ranges['call']:
                        ranges['call']['stop'] = cell.row - 1
                    elif 'stop' in ranges['call'] and cell_value == 'Strike' and 'start' not in ranges['put']:
                        ranges['put']['start'] = cell.row + 1
                    elif 'start' in ranges['put'] and cell_value == 'TOTALS':
                        ranges['put']['stop'] = cell.row - 1
                        break

            # Validate ranges
            if (ranges['call'].get('start') and ranges['call'].get('stop') and
                    ranges['put'].get('start') and ranges['put'].get('stop')):
                return ranges

            return None

        except Exception as e:
            logger.error(f"Ошибка поиска диапазонов: {e}")
            return None

    def _extract_data(self, ws, range_info: Dict) -> Dict:
        """Extract data from range"""
        data = {'strike': [], 'at_close': [], 'change': []}

        try:
            start_row = range_info.get('start')
            stop_row = range_info.get('stop')

            if not start_row or not stop_row:
                return data

            for row in range(start_row, stop_row + 1):
                try:
                    strike_val = ws[f'A{row}'].value
                    at_close_val = ws[f'I{row}'].value
                    change_val = ws[f'J{row}'].value

                    if strike_val is None or at_close_val is None or change_val is None:
                        continue

                    # Clean and convert values
                    strike_clean = str(strike_val).replace(",", "").replace("'", "")
                    at_close_clean = str(at_close_val).replace(",", "")
                    change_clean = str(change_val).replace(",", "")

                    data['strike'].append(int(float(strike_clean)))
                    data['at_close'].append(int(float(at_close_clean)))
                    data['change'].append(int(float(change_clean)))

                except (ValueError, TypeError, AttributeError):
                    continue

        except Exception as e:
            logger.error(f"Ошибка извлечения данных: {e}")

        return data

    def _calculate_sip(self, call_data: Dict, put_data: Dict, close_price: float,
                       coefficient: float, currency: str) -> Dict:
        """Calculate SIP metrics"""
        try:
            call_df = pd.DataFrame(call_data)
            put_df = pd.DataFrame(put_data)

            if call_df.empty or put_df.empty:
                return {}

            call_less = call_df[call_df['strike'] < close_price]
            put_more = put_df[put_df['strike'] > close_price]

            if currency in ["EUR", "GBP", "AUD", "XAU", "XAG"]:
                # Direct calculation
                up_level = (call_df['strike'] * call_df['at_close']).sum() / call_df['at_close'].sum() / coefficient
                down_level = (put_df['strike'] * put_df['at_close']).sum() / put_df['at_close'].sum() / coefficient

                up_balance = 0
                if len(put_more) > 0:
                    up_balance = (put_more['strike'] * put_more['at_close']).sum() / put_more[
                        'at_close'].sum() / coefficient

                down_balance = 0
                if len(call_less) > 0:
                    down_balance = (call_less['strike'] * call_less['at_close']).sum() / call_less[
                        'at_close'].sum() / coefficient

                red_balance = 0
                combined_volume = call_less['at_close'].sum() + put_more['at_close'].sum()
                if combined_volume > 0:
                    red_balance = ((call_less['strike'] * call_less['at_close']).sum() +
                                   (put_more['strike'] * put_more['at_close']).sum()) / combined_volume / coefficient
            else:
                # Inverted calculation for CAD, JPY
                up_level = (1 / (
                            (put_df['strike'] * put_df['at_close']).sum() / put_df['at_close'].sum())) * coefficient
                down_level = (1 / (
                            (call_df['strike'] * call_df['at_close']).sum() / call_df['at_close'].sum())) * coefficient

                up_balance = 0
                if len(call_less) > 0:
                    up_balance = (1 / ((call_less['strike'] * call_less['at_close']).sum() / call_less[
                        'at_close'].sum())) * coefficient

                down_balance = 0
                if len(put_more) > 0:
                    down_balance = (1 / ((put_more['strike'] * put_more['at_close']).sum() / put_more[
                        'at_close'].sum())) * coefficient

                red_balance = 0
                combined_volume = call_less['at_close'].sum() + put_more['at_close'].sum()
                if combined_volume > 0:
                    combined_weighted = ((call_less['strike'] * call_less['at_close']).sum() +
                                         (put_more['strike'] * put_more['at_close']).sum()) / combined_volume
                    red_balance = (1 / combined_weighted) * coefficient

            return {
                'up_level': float(up_level),
                'down_level': float(down_level),
                'up_balance_level': float(up_balance),
                'down_balance_level': float(down_balance),
                'red_balance_level': float(red_balance)
            }

        except Exception as e:
            logger.error(f"Ошибка расчета SIP: {e}")
            return {}

    def _calculate_fob(self, call_data: Dict, put_data: Dict, close_price: float) -> Dict:
        """Calculate FOB metrics"""
        try:
            call_df = pd.DataFrame(call_data)
            put_df = pd.DataFrame(put_data)

            if call_df.empty or put_df.empty:
                return {}

            call_less = call_df[call_df['strike'] < close_price]
            put_more = put_df[put_df['strike'] > close_price]
            call_more = call_df[call_df['strike'] > close_price]
            put_less = put_df[put_df['strike'] < close_price]

            return {
                'opt_in_money_call_i': int(call_less['at_close'].sum()),
                'opt_in_money_call_j': int(call_less['change'].sum()),
                'opt_in_money_put_i': int(put_more['at_close'].sum()),
                'opt_in_money_put_j': int(put_more['change'].sum()),
                'opt_without_money_call_i': int(call_more['at_close'].sum()),
                'opt_without_money_call_j': int(call_more['change'].sum()),
                'opt_without_money_put_i': int(put_less['at_close'].sum()),
                'opt_without_money_put_j': int(put_less['change'].sum())
            }

        except Exception as e:
            logger.error(f"Ошибка расчета FOB: {e}")
            return {}

    def _format_strikes(self, call_data: Dict, put_data: Dict, coefficient: float, currency: str) -> Dict:
        """Format strike data"""
        try:
            result = {"calls": [], "puts": []}

            max_len = max(len(call_data['strike']), len(put_data['strike']))

            for i in range(max_len):
                if i < len(call_data['strike']):
                    if currency in ["EUR", "GBP", "AUD", "XAU", "XAG"]:
                        price = call_data['strike'][i] / coefficient
                    else:
                        price = (1 / call_data['strike'][i]) * coefficient

                    result["calls"].append({
                        "price": float(price),
                        "strike": call_data['at_close'][i],
                        "delta": call_data['change'][i]
                    })

                if i < len(put_data['strike']):
                    if currency in ["EUR", "GBP", "AUD", "XAU", "XAG"]:
                        price = put_data['strike'][i] / coefficient
                    else:
                        price = (1 / put_data['strike'][i]) * coefficient

                    result["puts"].append({
                        "price": float(price),
                        "strike": put_data['at_close'][i],
                        "delta": put_data['change'][i]
                    })

            return result

        except Exception as e:
            logger.error(f"Ошибка форматирования страйков: {e}")
            return {"calls": [], "puts": []}

    def _save_json(self, data: Dict, currency: str):
        """Save data to JSON file"""
        try:
            filename = f"FOB_{currency}_{datetime.date.today().isoformat()}.json"
            output_path = self.output_path / filename

            # Convert numpy types to native Python types
            def convert_types(obj):
                if isinstance(obj, (np.int64, np.int32)):
                    return int(obj)
                elif isinstance(obj, (np.float64, np.float32)):
                    return float(obj)
                elif isinstance(obj, dict):
                    return {key: convert_types(value) for key, value in obj.items()}
                elif isinstance(obj, list):
                    return [convert_types(item) for item in obj]
                else:
                    return obj

            clean_data = convert_types(data)

            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(clean_data, f, indent=2, ensure_ascii=False)

            logger.info(f"💾 {currency} → {filename}")

        except Exception as e:
            logger.error(f"❌ Ошибка сохранения {currency}: {e}")

    def process_cftc_data(self) -> bool:
        """Process CFTC data"""
        try:
            cftc_files = list(Path.cwd().glob("FinFut*.xls"))
            if not cftc_files:
                logger.warning("⚠️ CFTC файл не найден")
                return False

            cftc_file = cftc_files[0]
            logger.info(f"📊 Обработка CFTC: {cftc_file.name}")

            workbook = xlrd.open_workbook(str(cftc_file))
            data = []

            currency_map = {
                'CANADIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE': 'CAD',
                'SWISS FRANC - CHICAGO MERCANTILE EXCHANGE': 'CHF',
                'BRITISH POUND STERLING - CHICAGO MERCANTILE EXCHANGE': 'GBP',
                'JAPANESE YEN - CHICAGO MERCANTILE EXCHANGE': 'JPY',
                'EURO FX - CHICAGO MERCANTILE EXCHANGE': 'EUR',
                'AUSTRALIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE': 'AUD'
            }

            for sheet_name in workbook.sheet_names():
                try:
                    sheet = workbook.sheet_by_name(sheet_name)

                    for row_idx in range(1, sheet.nrows):
                        try:
                            row = sheet.row(row_idx)
                            if len(row) < 10:
                                continue

                            market_desc = str(row[0].value).strip()
                            currency = None

                            for key, curr in currency_map.items():
                                if key in market_desc:
                                    currency = curr
                                    break

                            if currency:
                                date_val = int(row[1].value)
                                date_str = f"20{date_val:06d}"
                                formatted_date = f"{date_str[:4]}.{date_str[4:6]}.{date_str[6:]}"

                                long_pos = int(float(row[8].value)) if row[8].value else 0
                                short_pos = int(float(row[9].value)) if row[9].value else 0

                                data.append({
                                    'currency': currency,
                                    'date': formatted_date,
                                    'long': long_pos,
                                    'short': short_pos
                                })

                        except Exception:
                            continue
                except Exception:
                    continue

            if data:
                filename = "FinFut.json"
                output_path = self.output_path / filename

                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)

                logger.info(f"✅ CFTC: {len(data)} записей → {filename}")

                # Cleanup
                cftc_file.unlink()
                return True
            else:
                logger.error("❌ Нет данных CFTC")
                return False

        except Exception as e:
            logger.error(f"❌ Ошибка обработки CFTC: {e}")
            return False

    def cleanup(self):
        """Clean up temporary files"""
        try:
            # Remove temporary files
            temp_files = ["fut_fin.zip", "fut_fin_2025.zip", "FinFutYY.xls"]
            for temp_file in temp_files:
                temp_path = Path(temp_file)
                if temp_path.exists():
                    temp_path.unlink()
                    logger.info(f"🗑️ Удален: {temp_file}")

            logger.info("✅ Очистка завершена")

        except Exception as e:
            logger.warning(f"⚠️ Ошибка очистки: {e}")


def format_duration(seconds: float) -> str:
    """Format duration"""
    if seconds < 60:
        return f"{seconds:.1f}с"
    elif seconds < 3600:
        return f"{seconds / 60:.1f}м"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        return f"{hours}ч {minutes}м"


def main():
    """Main function"""
    start_time = time.time()

    try:
        print("\n" + "🚀 ОБРАБОТЧИК ОПЦИОННЫХ ДАННЫХ")
        print("=" * 60)

        # Initialize processor
        processor = SimpleOptionsProcessor()
        processor.show_info()

        # Step 1: Download CME files
        step_time = time.time()
        if not processor.download_cme_files():
            logger.error("❌ Ошибка загрузки CME")
            return False
        cme_time = time.time() - step_time

        # Step 2: Download CFTC (optional)
        step_time = time.time()
        print("\n📊 Загрузка данных CFTC")
        choice = input("Загрузить CFTC данные? (y/n, по умолчанию y): ").strip().lower()

        cftc_success = False
        if not choice or choice.startswith('y'):
            cftc_success = processor.download_cftc_simple()
        else:
            logger.info("⏭️ Пропуск CFTC данных")

        cftc_time = time.time() - step_time

        # Step 3: Convert files
        step_time = time.time()
        if not processor.convert_files():
            logger.error("❌ Ошибка конвертации")
            return False
        convert_time = time.time() - step_time

        # Step 4: Get prices
        step_time = time.time()
        logger.info("💹 Получение цен...")

        prices = processor.get_close_prices()
        if not prices:
            print("\n🔄 MT5 недоступен. Выберите альтернативу:")
            print("1. ✏️  Ввести цены вручную")
            print("2. 📊 Использовать примерные цены")

            while True:
                try:
                    choice = input("\nВыберите (1-2, по умолчанию 2): ").strip()
                    if not choice or choice == "2":
                        prices = processor.get_alternative_prices()
                        break
                    elif choice == "1":
                        prices = processor.get_manual_prices()
                        if prices:
                            break
                    else:
                        print("❌ Неверный выбор")
                        continue
                except KeyboardInterrupt:
                    logger.info("🛑 Отмена")
                    return False

        if not prices:
            logger.error("❌ Нет цен для обработки")
            return False

        prices_time = time.time() - step_time

        # Step 5: Process currencies
        step_time = time.time()
        currencies = list(processor.currencies.keys())
        processed = 0

        for i, currency in enumerate(currencies):
            if processor.process_currency(currency, prices[i]):
                processed += 1

        process_time = time.time() - step_time

        # Step 6: Process CFTC if downloaded
        if cftc_success:
            processor.process_cftc_data()

        # Results
        total_time = time.time() - start_time

        print(f"\n🎉 ЗАВЕРШЕНО!")
        print("=" * 40)
        print(f"⏱️ Общее время: {format_duration(total_time)}")
        print(f"📥 CME: {format_duration(cme_time)}")
        print(f"📊 CFTC: {format_duration(cftc_time)}")
        print(f"🔄 Конвертация: {format_duration(convert_time)}")
        print(f"💹 Цены: {format_duration(prices_time)}")
        print(f"⚙️ Обработка: {format_duration(process_time)}")
        print(f"✅ Успешно: {processed}/{len(currencies)}")

        processor.show_info()
        processor.cleanup()

        print(f"\n💾 Результаты: {processor.output_path}")

        return True

    except KeyboardInterrupt:
        elapsed = time.time() - start_time
        print(f"\n🛑 Прервано после {format_duration(elapsed)}")
        return False
    except Exception as e:
        elapsed = time.time() - start_time
        logger.error(f"❌ Критическая ошибка после {format_duration(elapsed)}: {e}")
        return False


if __name__ == "__main__":
    main()